from Excel_operation.services import PersistServices
from Excel_operation.productinfo import Product
import openpyxl

'''
text
csv
json
excel
database -->
'''

def create_fresh_workbook_getsheet():
    workbook = openpyxl.Workbook()
    try:
        workbook.remove(workbook['prodinfo'])# always new work
    # workbook = openpyxl.load_workbook('D:\\python_work\\temp\\product.xlsx')    # if file present--same use kr
    except:
        pass
    return workbook


def add_header(sheet):
    sheet.cell(1, 1, "PRODUCT_ID")
    sheet.cell(1, 2, "PRODUCT_NAME")
    sheet.cell(1, 3, "PRODUCT_QTY")
    sheet.cell(1, 4, "PRODUCT_PRICE")
    sheet.cell(1, 5, "PRODUCT_CATEGORY")
    return sheet

FILE_PATH = 'C:\\My Data\\Flask\\Excel_operation\\product.xlsx'


def write_data_into_excel(listOfProducts):    # excel --workbook --> sheets --> rows/cols --> cell -- data
    ''' THIS METHOD IS GOING TO WRITE-- LIST OF PRODUCTS INTO FRESH EXCELSHEET..!'''
    workbook = create_fresh_workbook_getsheet()
    sheet = workbook.create_sheet('prodinfo')
    sheet = add_header(sheet)
    row = 2
    for prod in listOfProducts:
        sheet.cell(row, 1, prod.prodId)
        sheet.cell(row, 2, prod.prodName)
        sheet.cell(row, 3, prod.prodQty)
        sheet.cell(row, 4, prod.prodPrice)
        sheet.cell(row, 5, prod.prodCategory)
        row = row + 1

    workbook.save(FILE_PATH)
    print('Data Written into Excel File..!')



def read_data_from_excel():
    ''' THIS METHOD IS GOING READ THE CONTENTS FROM EXCEL-- GOING TO CREATE LIST OF PRODUCTS...'''

    prodlist = []
    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
        sheet = workbook['prodinfo']
        maxrows = sheet.max_row+1
        #max_cols = sheet.max_column

        for row in range(2,maxrows):
            pid = sheet.cell(row, 1).value  # -->       #2,1  -- id
            pnm = sheet.cell(row, 2).value  # -->       #2,1  -->name
            pqty = sheet.cell(row, 3).value  # -->       #2,1 --> qty
            prc = sheet.cell(row, 4).value  # -->       #2,1 -->price
            cat = sheet.cell(row, 5).value  # -->       #2,1 -> category
            prodlist.append(Product(pid,pnm,prc,pqty,cat))
    except:
        print("No file with given Name")
    return prodlist


excelProductList = []  # this list be initialized --> when -- > u create an instance of ExcelserviceImpl-->
# as long as u dont have instance of Excelservice--u r not allowed to any of the service..


class ExcelServiceImpl(PersistServices):

    def __init__(self):
        global excelProductList
        excelProductList = read_data_from_excel()

    def add_product(self, prod):
        if type(prod) == Product:
            if type(prod.prodId)==int and prod.prodId > 0:
                if self.get_product(prod.prodId) and self.get_product(prod.prodName) and self.get_product(prod.prodPrice) and self.get_product(prod.prodCategory):  #    product or None
                    print('Given Product Already Exist into Excel sheet --Duplicate Product')
                else:
                    excelProductList.append(prod)
                    write_data_into_excel(excelProductList) # sync up -- memory data with--excelsheet..
                    print('Product Added Successfully....!')
            else:
                print('Invalid Product Id...Cannot add product into inventory')
        else:
            print('Invalid Product Type.!')

    def get_product(self, prId):  # product or None
        for prod in self.get_all_products():
            if prod.prodId == prId:
                return prod

    def get_all_products(self): # read from excelsheet-->
        excelProductList = read_data_from_excel()
        return excelProductList

   
    def update_product(self, prId, prod):
        for eprod in self.get_all_products():
            if eprod.prodId == prId:
                if eprod:
                    eprod.prodName  = prod.prodName
                    eprod.prodQty = prod.prodQty
                    eprod.prodPrice = prod.prodPrice
                    eprod.prodCategory = prod.prodCategory
                    # workbook.remove(workbook['empdata'])
                    # self.add_product(eprod)
        for prod in self.get_all_products():
            self.add_product(prod)
            print('Product Record Updated Successfully...!')
        # write_data_into_excel(excelProductList) #sync up memory data with excelsheet
            return

        # print('Cannot update update product.. no product found with given Id..!')

   
    def remove_product(self, prId):
        for prod in self.get_all_products():
            if prod.prodId == prId:
                excelProductList.remove(prod)
                print('Product removed from Excel.!')
                write_data_into_excel(excelProductList)# sync up memory data with -- excelsheet..
                return
        print('No Product with Given Id..so cannot delete..!')

    def get_products_by_category(self, cate):
        prodCategoryList = []
        for prod in self.get_all_products():
            if prod.prodCategory == cate:
                prodCategoryList.append(prod)
        if prodCategoryList:
            return prodCategoryList
        print('No products with given Category...!')
   
    def get_max_product_price(self):
        if self.get_all_products():
            #excelProductList.sort(key = lambda prod : prod.prodPrice)
            #return excelProductList[-1]
            return sorted(self.get_all_products(),key=lambda prod : prod.prodPrice)[-1]



if __name__ == '__main__':
    service = ExcelServiceImpl()
    # print(service.get_max_product_price())

    # import sys
    # sys.exit(0)
    prod8 = Product(pid=101,pnm='SDR',prc=809809,pqty=30,pcat='B+')

    service.update_product(101,prod8)

    import sys
    sys.exit(0)
    prod1 = Product(pid=101,pnm='AAAA1',prc=29325.3,pqty=33,pcat='A')
    prod2 = Product(pid=102, pnm='AAAA2', prc=24932.3, pqty=53, pcat='A')
    prod3 = Product(pid=103, pnm='AAAA3', prc=62932.3, pqty=23, pcat='C')
    prod4 = Product(pid=104, pnm='AAAA4', prc=25932.3, pqty=63, pcat='D')
    prod5 = Product(pid=105, pnm='AAAA5', prc=42932.3, pqty=73, pcat='A')

    products = [prod1,prod2,prod3,prod4,prod5]

    for prod in products:
        service.add_product(prod)


    # service.remove_product(102)


